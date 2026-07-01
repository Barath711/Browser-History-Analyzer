#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Browser History Analyzer  v2.0
================================
PyQt6 GUI for Edge/Chrome browsing & download history.
  • Preview history + downloads on-screen before export
  • Correlate downloads with browsing history
  • All timestamps displayed in UTC
  • iOS-style dark/light theme (MultiOSINTv11 UI reference)

Requirements:
    pip install PyQt6 openpyxl
"""
# ── Circular animated splash screen (PyQt6) ─────────────────────────────────
import sys as _sys, os as _os

from PyQt6.QtWidgets import QApplication as _QApp, QWidget as _QW
from PyQt6.QtCore    import Qt as _Qt, QTimer as _QTimer, QRect as _QRect
from PyQt6.QtGui     import (QPainter as _QP, QColor as _QCol, QPen as _QPen,
                              QFont as _QFont, QBrush as _QBrush)

_app_qt = _QApp.instance() or _QApp(_sys.argv)
_app_qt.setApplicationName("Browsing History")


class _CircularSplash(_QW):
    """Animated circular loading splash (GeisielMelo-style)."""
    def __init__(self):
        super().__init__()
        self.setWindowFlags(
            _Qt.WindowType.FramelessWindowHint |
            _Qt.WindowType.WindowStaysOnTopHint |
            _Qt.WindowType.Tool)
        self.setAttribute(_Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(340, 340)
        sg = _app_qt.primaryScreen().availableGeometry()
        self.move((sg.width() - 340) // 2, (sg.height() - 340) // 2)
        self._angle  = 0
        self._pct    = 0
        self._status = "Initializing..."
        self._timer  = _QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(16)

    def set_status(self, msg: str, bump: int = 15):
        self._status = msg
        self._pct = min(self._pct + bump, 95)
        _app_qt.processEvents()

    def _tick(self):
        self._angle = (self._angle + 3) % 360
        self.update()

    def paintEvent(self, _e):
        p = _QP(self)
        p.setRenderHint(_QP.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        cx, cy = W // 2, H // 2
        R = 140
        # Dark background circle
        p.setBrush(_QBrush(_QCol("#1C1C1E")))
        p.setPen(_Qt.PenStyle.NoPen)
        p.drawEllipse(cx - R, cy - R, R * 2, R * 2)
        # Outer track ring
        m = 10
        trk = _QPen(_QCol("#3A3A3C"), 9, _Qt.PenStyle.SolidLine, _Qt.PenCapStyle.RoundCap)
        p.setPen(trk); p.setBrush(_Qt.BrushStyle.NoBrush)
        p.drawArc(cx-R+m, cy-R+m, (R-m)*2, (R-m)*2, 0, 360*16)
        # Animated blue arc
        arc = _QPen(_QCol("#0A84FF"), 9, _Qt.PenStyle.SolidLine, _Qt.PenCapStyle.RoundCap)
        p.setPen(arc)
        p.drawArc(cx-R+m, cy-R+m, (R-m)*2, (R-m)*2, (90 - self._angle)*16, -130*16)
        # App title
        p.setPen(_QCol("#F2F2F7"))
        p.setFont(_QFont("Segoe UI", 14, _QFont.Weight.Bold))
        p.drawText(_QRect(0, cy-62, W, 30), _Qt.AlignmentFlag.AlignHCenter, "Browsing History")
        # Subtitle
        p.setPen(_QCol("#636366"))
        p.setFont(_QFont("Segoe UI", 9))
        p.drawText(_QRect(0, cy-30, W, 22), _Qt.AlignmentFlag.AlignHCenter, "Analyzer")
        # Version badge
        vr = _QRect(cx - 35, cy - 5, 70, 22)
        p.setBrush(_QBrush(_QCol("#2C2C2E"))); p.setPen(_Qt.PenStyle.NoPen)
        p.drawRoundedRect(vr, 11, 11)
        p.setPen(_QCol("#5AC8FA")); p.setFont(_QFont("Segoe UI", 8))
        p.drawText(vr, _Qt.AlignmentFlag.AlignCenter, "v 2.0")
        # Percent
        p.setPen(_QCol("#5AC8FA"))
        p.setFont(_QFont("Segoe UI", 22, _QFont.Weight.Bold))
        p.drawText(_QRect(0, cy+22, W, 44), _Qt.AlignmentFlag.AlignHCenter, f"{self._pct}%")
        # Status text
        p.setPen(_QCol("#8E8E93")); p.setFont(_QFont("Segoe UI", 9))
        p.drawText(_QRect(0, cy+68, W, 22), _Qt.AlignmentFlag.AlignHCenter, self._status)
        p.end()


_splash_win = _CircularSplash()
_splash_win.show()
_app_qt.processEvents()

def _splash_status(msg: str, bump: int = 15):
    _splash_win.set_status(msg, bump)

_splash_status("Loading libraries...", 10)



import os, sys, shutil, sqlite3, tempfile, traceback
import datetime as dt
from collections import defaultdict
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QCheckBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QFrame, QFileDialog, QMessageBox,
    QSplitter, QProgressBar, QAbstractItemView, QTabWidget, QSizePolicy)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QIcon

# ─── Constants ──────────────────────────────────────────────────────────────
WEBKIT_EPOCH  = dt.datetime(1601, 1, 1, tzinfo=dt.timezone.utc)
ERROR_LOG     = os.path.join(os.environ.get("TEMP", "."), "history_export_error.log")
_SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PREVIEW_LIMIT = 5_000   # max rows shown in table preview

# ─── iOS-style theme (matches MultiOSINTv11) ────────────────────────────────
THEME = {
    "dark": {
        "bg":     "#1C1C1E", "bg2":    "#2C2C2E", "bg3":    "#3A3A3C",
        "bg4":    "#48484A", "accent": "#0A84FF", "green":  "#30D158",
        "red":    "#FF453A", "orange": "#FF9F0A", "yellow": "#FFD60A",
        "fg":     "#FFFFFF", "fg2":    "#EBEBF5", "fg3":    "#8E8E93",
        "border": "#3A3A3C", "card":   "#2C2C2E",
    },
    "light": {
        "bg":     "#F2F2F7", "bg2":    "#FFFFFF", "bg3":    "#E5E5EA",
        "bg4":    "#D1D1D6", "accent": "#007AFF", "green":  "#34C759",
        "red":    "#FF3B30", "orange": "#FF9500", "yellow": "#FFCC00",
        "fg":     "#000000", "fg2":    "#3C3C43", "fg3":    "#6C6C70",
        "border": "#C6C6C8", "card":   "#FFFFFF",
    }
}

def build_stylesheet(mode: str = "dark") -> str:
    t = THEME[mode]
    bg  = t["bg"];  bg2 = t["bg2"]; bg3 = t["bg3"]; bg4 = t["bg4"]
    fg  = t["fg"];  fg3 = t["fg3"]; acc = t["accent"]
    bdr = t["border"]; grn = t["green"]; red = t["red"]
    acc2 = "#3395FF" if mode == "dark" else "#1A8AFF"
    return (
        f"QMainWindow,QWidget{{background:{bg};color:{fg};"
        f"font-family:'Segoe UI',-apple-system,sans-serif;font-size:13px}}"
        f"QLabel{{color:{fg};background:transparent}}"
        f"QLabel[dim='true']{{color:{fg3}}}"
        f"QLabel[accent='true']{{color:{acc}}}"
        f"QPushButton{{background:{bg3};color:{fg};border:none;border-radius:10px;"
        f"padding:7px 16px;font-size:13px;font-weight:500}}"
        f"QPushButton:hover{{background:{acc};color:#fff}}"
        f"QPushButton:pressed{{background:{bg4}}}"
        f"QPushButton[accent='true']{{background:{acc};color:#fff;font-weight:600}}"
        f"QPushButton[accent='true']:hover{{background:{acc2}}}"
        f"QPushButton[success='true']{{background:{grn};color:#fff}}"
        f"QPushButton[success='true']:hover{{background:#2ec14e;color:#fff}}"
        f"QPushButton[danger='true']{{background:{red};color:#fff}}"
        f"QPushButton[themebtn='true']{{background:{bg3};color:{fg};"
        f"border:1.5px solid {bdr};border-radius:10px;font-size:16px;"
        f"min-width:36px;min-height:28px;padding:2px 8px}}"
        f"QPushButton[themebtn='true']:hover{{background:{acc};color:#fff}}"
        f"QPushButton:disabled{{background:{bg2};color:{fg3}}}"
        f"QLineEdit{{background:{bg2};color:{fg};border:1.5px solid {bdr};"
        f"border-radius:10px;padding:6px 10px;selection-background-color:{acc}}}"
        f"QLineEdit:focus{{border-color:{acc}}}"
        f"QCheckBox{{color:{fg};spacing:8px}}"
        f"QCheckBox::indicator{{width:18px;height:18px;border-radius:5px;"
        f"border:2px solid {bdr};background:{bg2}}}"
        f"QCheckBox::indicator:checked{{background:{acc};border-color:{acc}}}"
        f"QTableWidget{{background:{bg};color:{fg};border:none;"
        f"gridline-color:{bdr};alternate-background-color:{bg2}}}"
        f"QTableWidget::item{{padding:4px 8px;border:none}}"
        f"QTableWidget::item:selected{{background:{acc};color:#fff}}"
        f"QHeaderView::section{{background:{bg2};color:{fg3};border:none;"
        f"border-bottom:1px solid {bdr};padding:6px 8px;font-weight:600;font-size:12px}}"
        f"QScrollBar:vertical{{background:{bg2};width:6px;border-radius:3px;margin:0}}"
        f"QScrollBar::handle:vertical{{background:{bg4};border-radius:3px;min-height:30px}}"
        f"QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{{height:0}}"
        f"QScrollBar:horizontal{{background:{bg2};height:6px;border-radius:3px}}"
        f"QScrollBar::handle:horizontal{{background:{bg4};border-radius:3px;min-width:30px}}"
        f"QScrollBar::add-line:horizontal,QScrollBar::sub-line:horizontal{{width:0}}"
        f"QProgressBar{{background:{bg3};border:none;border-radius:5px;"
        f"text-align:center;color:transparent}}"
        f"QProgressBar::chunk{{background:{acc};border-radius:5px}}"
        f"QSplitter::handle{{background:{bdr}}}"
        f"QFrame[sidebar='true']{{background:{bg2};border-right:1px solid {bdr}}}"
        f"QFrame[hdr='true']{{background:{bg2};border-bottom:1px solid {bdr}}}"
        f"QFrame[sep_line='true']{{background:{bdr};max-height:1px;border:none}}"
        f"QTabWidget::pane{{background:{bg};border:none;border-top:1px solid {bdr}}}"
        f"QTabBar{{background:{bg2};border-bottom:1px solid {bdr}}}"
        f"QTabBar::tab{{background:transparent;color:{fg3};border:none;"
        f"padding:10px 18px;font-size:13px}}"
        f"QTabBar::tab:selected{{color:{acc};font-weight:600;border-bottom:2px solid {acc}}}"
        f"QTabBar::tab:hover{{color:{fg};background:{bg3}}}"
    )

# ─── UTC timestamp conversion ────────────────────────────────────────────────
def webkit_ts_to_utc(ts_micro: Optional[int]) -> str:
    """Chromium µs since 1601-01-01 UTC → UTC string."""
    if ts_micro is None:
        return ""
    try:
        val = int(ts_micro)
        if val <= 0:
            return ""
        utc_dt = WEBKIT_EPOCH + dt.timedelta(microseconds=val)
        return utc_dt.strftime("%Y-%m-%d  %H:%M:%S") + "  UTC"
    except Exception:
        return ""

# ─── DB helpers ──────────────────────────────────────────────────────────────
def log_error(e: Exception) -> None:
    try:
        with open(ERROR_LOG, "a", encoding="utf-8") as f:
            f.write("\n=== ERROR ===\n")
            f.write("Time: " + dt.datetime.utcnow().isoformat(timespec="seconds") + " UTC\n")
            f.write(traceback.format_exc() + "\n")
    except Exception:
        pass

def safe_copy_db(src_path: str) -> str:
    if not os.path.isfile(src_path):
        raise FileNotFoundError(f"History DB not found: {src_path}")
    tmp_dir  = tempfile.mkdtemp(prefix="history_export_")
    tmp_path = os.path.join(tmp_dir, "History_copy.sqlite")
    shutil.copy2(src_path, tmp_path)
    return tmp_path

def table_exists(conn: sqlite3.Connection, table: str) -> bool:
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
        return cur.fetchone() is not None
    except sqlite3.Error:
        return False

def get_table_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    if not table_exists(conn, table):
        return []
    try:
        cur = conn.cursor()
        cur.execute(f"PRAGMA table_info({table})")
        return [row[1] for row in cur.fetchall()]
    except Exception:
        return []

def _first_nonempty(row: tuple, col_idx: Dict[str, int], *names: str) -> str:
    for n in names:
        if n in col_idx:
            v = row[col_idx[n]]
            if v and str(v).strip():
                return str(v).strip()
    return ""

def _base_filename(*paths) -> str:
    for p in paths:
        if p and str(p).strip():
            try:
                return os.path.basename(str(p).strip())
            except Exception:
                continue
    return ""

_DL_STATE = {0: "In Progress", 1: "Complete", 2: "Cancelled",
             3: "Interrupted", 4: "Interrupted"}

# ─── Data loading ─────────────────────────────────────────────────────────────
def load_data(db_path: str, use_host_match: bool) -> Dict:
    """Load history + downloads from a Chrome/Edge History SQLite file."""
    result = {"history": [], "downloads": [], "error": None,
              "hist_count": 0, "dl_count": 0}
    tmp_path = None
    conn     = None
    try:
        tmp_path = safe_copy_db(db_path)
        try:
            conn = sqlite3.connect(f"file:{tmp_path}?mode=ro", uri=True)
        except Exception:
            conn = sqlite3.connect(tmp_path)

        # ── Build downloads lookup ─────────────────────────────────────
        useful_dl = [
            "id", "tab_url", "site_url", "referrer", "url",
            "target_path", "current_path", "suggested_path", "full_path",
            "start_time", "end_time", "bytes_received", "total_bytes",
            "state", "opened",
        ]
        dl_cols    = get_table_columns(conn, "downloads")
        select_dl  = [c for c in useful_dl if c in dl_cols]
        if "id" not in select_dl and dl_cols:
            select_dl.insert(0, "id")

        exact_url_to_files: Dict[str, Set[str]] = defaultdict(set)
        host_to_files:      Dict[str, Set[str]] = defaultdict(set)
        downloads_rows:     List[Dict]          = []

        if select_dl:
            cur    = conn.cursor()
            raw_dl = cur.execute(f"SELECT {', '.join(select_dl)} FROM downloads").fetchall()
            ci     = {name: i for i, name in enumerate(d[0] for d in cur.description)}

            # URL chains
            final_url_map: Dict[int, str]        = {}
            chain_map:     Dict[int, List[str]]  = defaultdict(list)
            if table_exists(conn, "downloads_url_chains"):
                try:
                    chain_rows = cur.execute(
                        "SELECT id, chain_index, url FROM downloads_url_chains "
                        "ORDER BY id, chain_index"
                    ).fetchall()
                    tmp_chains: Dict[int, List[Tuple[int, str]]] = defaultdict(list)
                    for did, idx, url in chain_rows:
                        tmp_chains[int(did)].append((int(idx), str(url) if url else ""))
                    for did, lst in tmp_chains.items():
                        lst.sort(key=lambda x: x[0])
                        urls = [u for _, u in lst if u]
                        chain_map[did] = urls
                        if urls:
                            final_url_map[did] = urls[-1]
                except Exception as e:
                    log_error(e)

            for row in raw_dl:
                try:
                    did       = int(row[ci["id"]]) if "id" in ci else 0
                    path_vals = [row[ci[c]] for c in
                                 ["target_path", "current_path", "suggested_path", "full_path"]
                                 if c in ci]
                    file_name = _base_filename(*path_vals)
                    final_url = final_url_map.get(did)
                    source_url = (final_url or
                                  _first_nonempty(row, ci, "tab_url", "site_url", "referrer", "url"))
                    url_chain  = " → ".join(chain_map.get(did, []))
                    start_utc  = webkit_ts_to_utc(row[ci["start_time"]] if "start_time" in ci else None)
                    end_utc    = webkit_ts_to_utc(row[ci["end_time"]]   if "end_time"   in ci else None)
                    raw_state  = row[ci["state"]] if "state" in ci else None
                    state_str  = _DL_STATE.get(raw_state, str(raw_state) if raw_state is not None else "")
                    tb         = row[ci["total_bytes"]] if "total_bytes" in ci else None
                    try:
                        tb = int(tb) if tb is not None else 0
                    except Exception:
                        tb = 0
                    if tb >= 1_073_741_824: size_str = f"{tb/1_073_741_824:.1f} GB"
                    elif tb >= 1_048_576:   size_str = f"{tb/1_048_576:.1f} MB"
                    elif tb >= 1024:        size_str = f"{tb/1024:.1f} KB"
                    elif tb > 0:            size_str = f"{tb} B"
                    else:                   size_str = ""

                    drec = {
                        "id":          did,
                        "file_name":   file_name or "",
                        "source_url":  source_url or "",
                        "final_url":   final_url  or "",
                        "url_chain":   url_chain,
                        "start_utc":   start_utc,
                        "end_utc":     end_utc,
                        "state":       state_str,
                        "size":        size_str,
                    }
                    downloads_rows.append(drec)
                    if source_url and file_name:
                        exact_url_to_files[source_url].add(file_name)
                        host = urlparse(source_url).netloc.lower()
                        if host:
                            host_to_files[host].add(file_name)
                except Exception as e:
                    log_error(e)

        result["downloads"] = downloads_rows
        result["dl_count"]  = len(downloads_rows)

        # ── History ────────────────────────────────────────────────────
        history_rows: List[Dict] = []
        if table_exists(conn, "urls"):
            cur      = conn.cursor()
            raw_hist = cur.execute(
                "SELECT url, title, last_visit_time FROM urls ORDER BY last_visit_time DESC"
            ).fetchall()
            for (url, title, last_visit_time) in raw_hist:
                utc_str  = webkit_ts_to_utc(last_visit_time)
                files: Set[str] = set()
                if url:
                    s_url = str(url).strip()
                    files |= exact_url_to_files.get(s_url, set())
                    if use_host_match:
                        host = urlparse(s_url).netloc.lower()
                        if host:
                            files |= host_to_files.get(host, set())
                files_str = "; ".join(sorted(files)) if files else ""
                history_rows.append({
                    "utc_time":        utc_str,
                    "url":             url   or "",
                    "title":           title or "",
                    "downloaded":      bool(files_str),
                    "downloaded_files": files_str,
                })

        result["history"]    = history_rows
        result["hist_count"] = len(history_rows)

    except Exception as e:
        log_error(e)
        result["error"] = str(e)
    finally:
        if conn:
            try:  conn.close()
            except Exception: pass
        if tmp_path:
            try:
                os.remove(tmp_path)
                os.rmdir(os.path.dirname(tmp_path))
            except Exception: pass
    return result

# ─── XLSX Export ─────────────────────────────────────────────────────────────
def export_xlsx(data: Dict, outdir: str, prefix: str) -> Tuple[str, int, int]:
    """Export history and downloads to xlsx.  All times in UTC."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl required. pip install openpyxl")
    wb       = Workbook()
    ws_hist  = wb.active
    ws_hist.title = "History"
    ws_dl    = wb.create_sheet("Downloads")

    ws_hist.append(["last_visit_time_utc", "url", "title", "downloaded", "downloaded_files"])
    for r in data["history"]:
        ws_hist.append([r["utc_time"], r["url"], r["title"],
                        "Yes" if r["downloaded"] else "No", r["downloaded_files"]])

    ws_dl.append(["id", "file_name", "source_url", "final_url", "url_chain",
                  "start_time_utc", "end_time_utc", "state", "size"])
    for r in data["downloads"]:
        ws_dl.append([r.get("id",""), r.get("file_name",""), r.get("source_url",""),
                      r.get("final_url",""), r.get("url_chain",""),
                      r.get("start_utc",""), r.get("end_utc",""),
                      r.get("state",""), r.get("size","")])

    for ws in [ws_hist, ws_dl]:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    os.makedirs(outdir, exist_ok=True)
    out_path = os.path.join(outdir, f"{prefix}_history.xlsx")
    wb.save(out_path)
    return out_path, len(data["history"]), len(data["downloads"])

# ─── Worker thread ────────────────────────────────────────────────────────────
class LoadWorker(QThread):
    finished = pyqtSignal(dict)
    error    = pyqtSignal(str)

    def __init__(self, db_path: str, use_host_match: bool):
        super().__init__()
        self.db_path        = db_path
        self.use_host_match = use_host_match

    def run(self):
        try:
            result = load_data(self.db_path, self.use_host_match)
            if result.get("error"):
                self.error.emit(result["error"])
            else:
                self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))

# ─── Main Window ──────────────────────────────────────────────────────────────
class HistoryApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self._dark:           bool            = True
        self._history_files:  List[str]       = []
        self._output_dir:     str             = (
            os.path.expanduser("~\\Desktop")
            if os.path.isdir(os.path.expanduser("~\\Desktop"))
            else os.getcwd()
        )
        self._loaded_data:   Optional[Dict]        = None
        self._all_hist_rows: List[Dict]            = []
        self._all_dl_rows:   List[Dict]            = []
        self._worker:        Optional[LoadWorker]  = None

        self.setWindowTitle("Browser History Analyzer")
        try:
            ico = os.path.join(_SCRIPT_DIR, "myicon.ico")
            if os.path.exists(ico):
                self.setWindowIcon(QIcon(ico))
        except Exception:
            pass
        self.setMinimumSize(1100, 700)
        self.resize(1350, 840)
        self._apply_theme()
        self._build_ui()
        _splash_win.hide()


    # ─── Theme ────────────────────────────────────────────────────────────
    def _apply_theme(self):
        self.setStyleSheet(build_stylesheet("dark" if self._dark else "light"))
        if hasattr(self, "_theme_btn"):
            self._theme_btn.setText("☀️" if self._dark else "🌙")
        if self._loaded_data and hasattr(self, "_hist_table"):
            self._render_history_table(self._all_hist_rows[:PREVIEW_LIMIT])
            self._render_downloads_table(self._all_dl_rows[:PREVIEW_LIMIT])

    def _toggle_theme(self):
        self._dark = not self._dark
        self._apply_theme()

    # ─── UI scaffold ──────────────────────────────────────────────────────
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._build_header())
        body = QSplitter(Qt.Orientation.Horizontal)
        body.setHandleWidth(1)
        body.addWidget(self._build_sidebar())
        body.addWidget(self._build_main())
        body.setSizes([290, 1060])
        root.addWidget(body, 1)
        root.addWidget(self._build_footer())

    # ─── Header ───────────────────────────────────────────────────────────
    def _build_header(self) -> QFrame:
        hdr = QFrame()
        hdr.setProperty("hdr", True)
        lay = QHBoxLayout(hdr)
        lay.setContentsMargins(18, 12, 18, 12)

        logo = QLabel("🌐  Browser History Analyzer")
        logo.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        logo.setProperty("accent", True)
        lay.addWidget(logo)

        sub = QLabel("Edge / Chrome History Explorer  ·  All timestamps in UTC")
        sub.setProperty("dim", True)
        sub.setFont(QFont("Segoe UI", 10))
        lay.addWidget(sub)
        lay.addStretch()

        self._theme_btn = QPushButton("☀️")
        self._theme_btn.setProperty("themebtn", True)
        self._theme_btn.setToolTip("Toggle dark / light theme")
        self._theme_btn.clicked.connect(self._toggle_theme)
        lay.addWidget(self._theme_btn)
        return hdr

    # ─── Sidebar ──────────────────────────────────────────────────────────
    def _build_sidebar(self) -> QFrame:
        side = QFrame()
        side.setProperty("sidebar", True)
        side.setMinimumWidth(255)
        side.setMaximumWidth(340)
        lay = QVBoxLayout(side)
        lay.setContentsMargins(14, 16, 14, 14)
        lay.setSpacing(10)

        # ── File section ──────────────────────────────────────────────
        _lbl = QLabel("History File")
        _lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        lay.addWidget(_lbl)

        sel_btn = QPushButton("📂  Select History File(s)…")
        sel_btn.setProperty("accent", "true")
        sel_btn.setMinimumHeight(38)
        sel_btn.clicked.connect(self._select_files)
        lay.addWidget(sel_btn)

        self._files_lbl = QLabel("No files selected")
        self._files_lbl.setProperty("dim", True)
        self._files_lbl.setFont(QFont("Segoe UI", 9))
        self._files_lbl.setWordWrap(True)
        lay.addWidget(self._files_lbl)

        self._sep(lay)

        # ── Output section ────────────────────────────────────────────
        _lbl2 = QLabel("Output")
        _lbl2.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        lay.addWidget(_lbl2)

        lbl_out = QLabel("Output folder:")
        lbl_out.setFont(QFont("Segoe UI", 9))
        lbl_out.setProperty("dim", True)
        lay.addWidget(lbl_out)

        out_row = QHBoxLayout()
        out_row.setSpacing(6)
        self._out_dir_edit = QLineEdit(self._output_dir)
        self._out_dir_edit.setPlaceholderText("Output folder…")
        out_row.addWidget(self._out_dir_edit)
        brw = QPushButton("…")
        brw.setFixedWidth(36)
        brw.clicked.connect(self._browse_outdir)
        out_row.addWidget(brw)
        lay.addLayout(out_row)

        lbl_sub = QLabel("Subject label (prefix, optional):")
        lbl_sub.setFont(QFont("Segoe UI", 9))
        lbl_sub.setProperty("dim", True)
        lay.addWidget(lbl_sub)

        self._label_edit = QLineEdit()
        self._label_edit.setPlaceholderText("CaseID / username / hostname…")
        lay.addWidget(self._label_edit)

        self._sep(lay)

        # ── Options section ────────────────────────────────────────────
        _lbl3 = QLabel("Options")
        _lbl3.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        lay.addWidget(_lbl3)

        self._host_match_cb = QCheckBox("Match downloads by hostname")
        self._host_match_cb.setChecked(True)
        self._host_match_cb.setFont(QFont("Segoe UI", 10))
        lay.addWidget(self._host_match_cb)

        hint = QLabel("Also correlates downloads to history entries by hostname,\nnot only by exact URL.")
        hint.setProperty("dim", True)
        hint.setFont(QFont("Segoe UI", 8))
        hint.setWordWrap(True)
        lay.addWidget(hint)

        self._sep(lay)

        # ── Actions ───────────────────────────────────────────────────
        _lbl4 = QLabel("Actions")
        _lbl4.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        lay.addWidget(_lbl4)

        self._preview_btn = QPushButton("🔍  Preview Data")
        self._preview_btn.setProperty("accent", "true")
        self._preview_btn.setMinimumHeight(40)
        self._preview_btn.clicked.connect(self._run_preview)
        lay.addWidget(self._preview_btn)

        self._export_btn = QPushButton("📊  Export to XLSX")
        self._export_btn.setProperty("success", "true")
        self._export_btn.setMinimumHeight(40)
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._run_export)
        lay.addWidget(self._export_btn)

        lay.addStretch()

        # ── Progress ──────────────────────────────────────────────────
        self._progress = QProgressBar()
        self._progress.setValue(0)
        self._progress.setTextVisible(False)
        self._progress.setFixedHeight(6)
        lay.addWidget(self._progress)

        self._status_lbl = QLabel("Select a History file to begin")
        self._status_lbl.setProperty("dim", True)
        self._status_lbl.setFont(QFont("Segoe UI", 9))
        self._status_lbl.setWordWrap(True)
        lay.addWidget(self._status_lbl)

        return side

    def _sep(self, layout):
        f = QFrame()
        f.setFrameShape(QFrame.Shape.HLine)
        f.setProperty("sep_line", True)
        layout.addWidget(f)

    # ─── Main panel ────────────────────────────────────────────────────────
    def _build_main(self) -> QWidget:
        w   = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._tabs = QTabWidget()
        self._tabs.setDocumentMode(True)
        lay.addWidget(self._tabs)

        self._tabs.addTab(self._build_history_tab(),   "📋  Browsing History")
        self._tabs.addTab(self._build_downloads_tab(), "⬇️  Downloads")

        return w

    def _build_history_tab(self) -> QWidget:
        w   = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Filter bar
        bar = QFrame()
        bar.setProperty("hdr", True)
        b_lay = QHBoxLayout(bar)
        b_lay.setContentsMargins(10, 7, 10, 7)
        b_lay.setSpacing(8)

        self._hist_count_lbl = QLabel("No data loaded  —  click Preview Data")
        self._hist_count_lbl.setProperty("dim", True)
        self._hist_count_lbl.setFont(QFont("Segoe UI", 9))
        b_lay.addWidget(self._hist_count_lbl)
        b_lay.addStretch()

        lbl_f = QLabel("Filter:")
        lbl_f.setFont(QFont("Segoe UI", 9))
        lbl_f.setProperty("dim", True)
        b_lay.addWidget(lbl_f)

        self._hist_filter = QLineEdit()
        self._hist_filter.setPlaceholderText("Search URL or title…")
        self._hist_filter.setFixedWidth(220)
        self._hist_filter.textChanged.connect(self._filter_history)
        b_lay.addWidget(self._hist_filter)

        self._dl_only_btn = QPushButton("⬇️ Downloads Only")
        self._dl_only_btn.setCheckable(True)
        self._dl_only_btn.setToolTip("Show only history rows that have associated downloads")
        self._dl_only_btn.toggled.connect(lambda _: self._filter_history(self._hist_filter.text()))
        b_lay.addWidget(self._dl_only_btn)

        lay.addWidget(bar)

        # Table
        self._hist_table = self._make_table(
            ["UTC Time", "URL", "Title", "Downloaded", "Downloaded Files"],
            [185, 360, 220, 100, 300])
        lay.addWidget(self._hist_table)
        return w

    def _build_downloads_tab(self) -> QWidget:
        w   = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Filter bar
        bar = QFrame()
        bar.setProperty("hdr", True)
        b_lay = QHBoxLayout(bar)
        b_lay.setContentsMargins(10, 7, 10, 7)
        b_lay.setSpacing(8)

        self._dl_count_lbl = QLabel("No data loaded")
        self._dl_count_lbl.setProperty("dim", True)
        self._dl_count_lbl.setFont(QFont("Segoe UI", 9))
        b_lay.addWidget(self._dl_count_lbl)
        b_lay.addStretch()

        lbl_f2 = QLabel("Filter:")
        lbl_f2.setFont(QFont("Segoe UI", 9))
        lbl_f2.setProperty("dim", True)
        b_lay.addWidget(lbl_f2)

        self._dl_filter = QLineEdit()
        self._dl_filter.setPlaceholderText("Search file name or URL…")
        self._dl_filter.setFixedWidth(220)
        self._dl_filter.textChanged.connect(self._filter_downloads)
        b_lay.addWidget(self._dl_filter)

        lay.addWidget(bar)

        # Table
        self._dl_table = self._make_table(
            ["ID", "File Name", "Start Time (UTC)", "End Time (UTC)",
             "State", "Size", "Source URL", "Final URL"],
            [45, 200, 190, 190, 100, 80, 300, 300])
        lay.addWidget(self._dl_table)
        return w

    def _make_table(self, headers: List[str], widths: List[int]) -> QTableWidget:
        t = QTableWidget()
        t.setAlternatingRowColors(True)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        t.horizontalHeader().setStretchLastSection(True)
        t.verticalHeader().setDefaultSectionSize(26)
        t.verticalHeader().setVisible(False)
        t.setColumnCount(len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.setShowGrid(True)
        for i, w in enumerate(widths):
            t.setColumnWidth(i, w)
        return t

    # ─── Footer ───────────────────────────────────────────────────────────
    def _build_footer(self) -> QFrame:
        f = QFrame()
        f.setProperty("hdr", True)
        f.setMaximumHeight(30)
        lay = QHBoxLayout(f)
        lay.setContentsMargins(14, 4, 14, 4)
        lbl = QLabel("Browser History Analyzer  ·  All times shown in UTC  ·  v2.0")
        lbl.setFont(QFont("Segoe UI", 9))
        lbl.setProperty("dim", True)
        lay.addWidget(lbl)
        lay.addStretch()
        return f

    # ─── File / dir selection ─────────────────────────────────────────────
    def _select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Edge / Chrome History file(s)", "", "All files (*.*)")
        if files:
            self._history_files = list(files)
            n = len(files)
            self._files_lbl.setText(
                os.path.basename(files[0]) if n == 1
                else f"{n} files selected")
            self._status_lbl.setText(
                f"{n} file(s) selected — click 🔍 Preview Data")
            self._loaded_data = None
            self._export_btn.setEnabled(False)
            self._hist_table.setRowCount(0)
            self._dl_table.setRowCount(0)
            self._hist_count_lbl.setText("No data loaded  —  click Preview Data")
            self._dl_count_lbl.setText("No data loaded")

    def _browse_outdir(self):
        path = QFileDialog.getExistingDirectory(
            self, "Select output folder", self._output_dir)
        if path:
            self._output_dir = path
            self._out_dir_edit.setText(path)

    # ─── Preview ──────────────────────────────────────────────────────────
    def _run_preview(self):
        if not self._history_files:
            QMessageBox.warning(self, "No File Selected",
                                "Please select one or more History files first.")
            return
        if self._worker and self._worker.isRunning():
            return
        db_path  = self._history_files[0]
        use_host = self._host_match_cb.isChecked()
        self._set_busy(True)
        self._status_lbl.setText(f"Loading  {os.path.basename(db_path)} …")
        self._progress.setRange(0, 0)   # indeterminate spinner
        self._worker = LoadWorker(db_path, use_host)
        self._worker.finished.connect(self._on_load_done)
        self._worker.error.connect(self._on_load_error)
        self._worker.start()

    def _on_load_done(self, data: Dict):
        self._loaded_data   = data
        self._all_hist_rows = data["history"]
        self._all_dl_rows   = data["downloads"]
        self._progress.setRange(0, 100)
        self._progress.setValue(100)
        self._set_busy(False)
        self._export_btn.setEnabled(True)

        hist_n  = data["hist_count"]
        dl_n    = data["dl_count"]
        dl_hits = sum(1 for h in data["history"] if h["downloaded"])
        shown_h = min(hist_n, PREVIEW_LIMIT)
        shown_d = min(dl_n,   PREVIEW_LIMIT)
        trunc_h = f"  (showing first {PREVIEW_LIMIT:,})" if hist_n > PREVIEW_LIMIT else ""
        trunc_d = f"  (showing first {PREVIEW_LIMIT:,})" if dl_n   > PREVIEW_LIMIT else ""

        self._hist_count_lbl.setText(
            f"{hist_n:,} history records{trunc_h}  ·  "
            f"{dl_hits:,} correlated with downloads")
        self._dl_count_lbl.setText(f"{dl_n:,} download records{trunc_d}")
        self._status_lbl.setText(
            f"Loaded  ·  {hist_n:,} history  ·  {dl_n:,} downloads  ·  "
            f"{dl_hits:,} correlated")

        self._render_history_table(data["history"][:PREVIEW_LIMIT])
        self._render_downloads_table(data["downloads"][:PREVIEW_LIMIT])

        if len(self._history_files) > 1:
            self._status_lbl.setText(
                self._status_lbl.text() +
                f"  ·  Previewing 1/{len(self._history_files)} files")

        QTimer.singleShot(3000, lambda: self._progress.setValue(0))

    def _on_load_error(self, err: str):
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        self._set_busy(False)
        self._status_lbl.setText(f"Error: {err[:80]}")
        QMessageBox.critical(
            self, "Load Error",
            f"Failed to load history database:\n\n{err}\n\nSee log:\n{ERROR_LOG}")

    def _set_busy(self, busy: bool):
        self._preview_btn.setEnabled(not busy)
        self._export_btn.setEnabled(not busy and self._loaded_data is not None)

    # ─── Table rendering ──────────────────────────────────────────────────
    def _render_history_table(self, rows: List[Dict]):
        t    = THEME["dark" if self._dark else "light"]
        green = QColor(t["green"])
        dim   = QColor(t["fg3"])
        self._hist_table.setUpdatesEnabled(False)
        self._hist_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            items = [
                row["utc_time"],
                row["url"],
                row["title"],
                "✅  Yes" if row["downloaded"] else "—",
                row["downloaded_files"],
            ]
            for j, val in enumerate(items):
                item = QTableWidgetItem(str(val))
                if j == 3:
                    item.setForeground(green if row["downloaded"] else dim)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._hist_table.setItem(i, j, item)
        self._hist_table.setUpdatesEnabled(True)

    def _render_downloads_table(self, rows: List[Dict]):
        self._dl_table.setUpdatesEnabled(False)
        self._dl_table.setRowCount(len(rows))
        t   = THEME["dark" if self._dark else "light"]
        grn = QColor(t["green"])
        org = QColor(t["orange"])
        red = QColor(t["red"])
        dim = QColor(t["fg3"])
        for i, row in enumerate(rows):
            vals = [
                str(row.get("id","")),
                row.get("file_name",""),
                row.get("start_utc",""),
                row.get("end_utc",""),
                row.get("state",""),
                row.get("size",""),
                row.get("source_url",""),
                row.get("final_url",""),
            ]
            for j, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if j == 4:   # State column coloring
                    if val == "Complete":
                        item.setForeground(grn)
                    elif val == "In Progress":
                        item.setForeground(org)
                    elif val in ("Cancelled", "Interrupted"):
                        item.setForeground(red)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if j == 5:   # Size column align right
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self._dl_table.setItem(i, j, item)
        self._dl_table.setUpdatesEnabled(True)

    # ─── Filtering ────────────────────────────────────────────────────────
    def _filter_history(self, text: str):
        if not self._all_hist_rows:
            return
        dl_only = self._dl_only_btn.isChecked()
        needle  = text.lower()
        rows = [
            r for r in self._all_hist_rows
            if (not needle or needle in r["url"].lower() or needle in r["title"].lower())
            and (not dl_only or r["downloaded"])
        ]
        self._render_history_table(rows[:PREVIEW_LIMIT])
        total = len(self._all_hist_rows)
        shown = len(rows)
        dl_hits = sum(1 for r in rows if r["downloaded"])
        self._hist_count_lbl.setText(
            f"Showing {shown:,} of {total:,}  ·  {dl_hits:,} with downloads")

    def _filter_downloads(self, text: str):
        if not self._all_dl_rows:
            return
        needle = text.lower()
        rows = [
            r for r in self._all_dl_rows
            if not needle
            or needle in r.get("file_name","").lower()
            or needle in r.get("source_url","").lower()
            or needle in r.get("final_url","").lower()
        ]
        self._render_downloads_table(rows[:PREVIEW_LIMIT])
        self._dl_count_lbl.setText(
            f"Showing {len(rows):,} of {len(self._all_dl_rows):,} download records")

    # ─── Export ───────────────────────────────────────────────────────────
    def _run_export(self):
        if not self._loaded_data:
            QMessageBox.warning(self, "No Data", "Run Preview first to load data.")
            return
        if not HAS_OPENPYXL:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required.\n\nInstall with:\n  pip install openpyxl")
            return

        outdir = self._out_dir_edit.text().strip() or self._output_dir
        label  = self._label_edit.text().strip()
        if not label and self._history_files:
            parent = os.path.basename(os.path.dirname(self._history_files[0]))
            label  = parent.replace(" ", "_") if parent else "history"
        prefix = (label or "history").replace(" ", "_")

        try:
            # Export first (previewed) file
            out_path, hist_n, dl_n = export_xlsx(self._loaded_data, outdir, prefix)
            msg = f"Saved:\n{out_path}\n\nHistory rows : {hist_n:,}\nDownloads rows: {dl_n:,}"

            # If multiple files were selected, export the rest silently
            if len(self._history_files) > 1:
                extra = []
                for extra_path in self._history_files[1:]:
                    try:
                        extra_data   = load_data(extra_path, self._host_match_cb.isChecked())
                        extra_label  = os.path.basename(os.path.dirname(extra_path)).replace(" ", "_") or prefix
                        ep, h, d     = export_xlsx(extra_data, outdir, extra_label)
                        extra.append(f"  {ep}  (H:{h:,} D:{d:,})")
                    except Exception as e:
                        extra.append(f"  ERROR {os.path.basename(extra_path)}: {e}")
                if extra:
                    msg += "\n\nAdditional files:\n" + "\n".join(extra)

            self._status_lbl.setText(f"Exported — {hist_n:,} history + {dl_n:,} downloads")
            QMessageBox.information(self, "Export Complete", msg)

        except Exception as e:
            log_error(e)
            QMessageBox.critical(
                self, "Export Error",
                f"Export failed:\n\n{e}\n\nSee log:\n{ERROR_LOG}")

# ─── Entry point ──────────────────────────────────────────────────────────────
def main():
    try:
        if sys.platform.startswith("win"):
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = QApplication(sys.argv)
    app.setApplicationName("Browser History Analyzer")
    win = HistoryApp()
    win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
